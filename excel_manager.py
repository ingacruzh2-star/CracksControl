from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from unidecode import unidecode
from rapidfuzz import process, fuzz


class ExcelManager:

    def __init__(self):
        self.archivo = Path("excel") / "Futbol adulto Malaquías Concha 2026.xlsx"

        if not self.archivo.exists():
            raise FileNotFoundError(
                f"No existe el archivo:\n{self.archivo}"
            )

        self.wb = load_workbook(self.archivo)
        self.ws = self.wb["Asistencia 2026"]


    # ----------------------------------------------------

    def normalizar(self, texto):

        if texto is None:
            return ""

        return (
            unidecode(str(texto))
            .lower()
            .strip()
        )


    # ----------------------------------------------------

    def buscar_columna_fecha(self):

        hoy = datetime.now().date()

        for col in range(5, self.ws.max_column + 1):

            valor = self.ws.cell(row=2, column=col).value

            if valor is None:
                continue

            if hasattr(valor, "date"):

                if valor.date() == hoy:
                    return col

            else:

                try:

                    fecha = datetime.strptime(
                        str(valor),
                        "%d-%m-%Y"
                    ).date()

                    if fecha == hoy:
                        return col

                except:
                    pass

        return None


    # ----------------------------------------------------

    def crear_columna_hoy(self):

        col = self.ws.max_column + 1

        self.ws.cell(
            row=2,
            column=col
        ).value = datetime.now()

        return col


    # ----------------------------------------------------

    def obtener_columna_fecha(self):

        col = self.buscar_columna_fecha()

        if col is not None:
            return col

        return self.crear_columna_hoy()


    # ----------------------------------------------------

    def jugadores(self):

        datos = {}

        for fila in range(5, self.ws.max_row + 1):

            nombre = self.ws.cell(
                row=fila,
                column=3
            ).value

            if nombre:

                datos[self.normalizar(nombre)] = {
                    "fila": fila,
                    "nombre": nombre
                }

        return datos


    # ----------------------------------------------------

    def buscar_jugador(self, texto):

        jugadores = self.jugadores()

        resultado = process.extractOne(
            self.normalizar(texto),
            jugadores.keys(),
            scorer=fuzz.WRatio
        )

        if resultado is None:
            return None

        nombre, score, _ = resultado

        if score < 75:
            return None

        return jugadores[nombre]


    # ----------------------------------------------------

    def registrar(self, nombre):

        jugador = self.buscar_jugador(nombre)

        if jugador is None:
            return False, "Jugador no encontrado"

        col = self.obtener_columna_fecha()

        fila = jugador["fila"]

        if self.ws.cell(
            row=fila,
            column=col
        ).value == 1:

            return False, "Ya estaba registrado"

        self.ws.cell(
            row=fila,
            column=col
        ).value = 1

        self.wb.save(self.archivo)

        return True, jugador["nombre"]


    # ----------------------------------------------------

    def eliminar(self, nombre):

        jugador = self.buscar_jugador(nombre)

        if jugador is None:
            return False

        col = self.obtener_columna_fecha()

        self.ws.cell(
            row=jugador["fila"],
            column=col
        ).value = None

        self.wb.save(self.archivo)

        return True


    # ----------------------------------------------------

    def guardar(self):

        self.wb.save(self.archivo)